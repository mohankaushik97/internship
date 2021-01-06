import datetime

import openpyxl

date_row = 1
emp_column = 2


def main(emp_name):
    update_attendance(emp_name)
    save_file()


def sheet():
    now = datetime.datetime.now()

    month = now.strftime("%b")
    year = now.strftime("%Y")
    sheet_name = month + " " + year

    return sheet_name


def date_and_time():
    now = datetime.datetime.now()

    month = now.strftime("%m")
    day = now.strftime("%d")
    day = day.lstrip("0")
    time = now.strftime("%X")
    year = now.strftime("%Y")

    date = datetime.datetime(int(year), int(month), int(day))

    # date = month + " " + day

    print(date)

    return date, time


def save_file():
    sheet_name = sheet()
    file = openpyxl.load_workbook('tracker.xlsx')
    current_sheet = file[sheet_name]

    file.save('tracker.xlsx')


def current():
    sheet_name = sheet()
    file = openpyxl.load_workbook('tracker.xlsx')
    current_sheet = file[sheet_name]
    # print(current_sheet)

    return current_sheet


def update_attendance(emp_name):
    current_sheet = current()

    emp_row = find_emp(emp_name)
    date, time = date_and_time()
    # print (date, time )

    for column in range(1, current_sheet.max_column):
        date_cell = current_sheet.cell(date_row, column)
        # print(date_cell.value)
        if date_cell.value == date:
            # print("date cell: " + str(date_cell.value))
            if current_sheet.cell(emp_row, column).value == "Present":
                cell_updater = current_sheet.cell(emp_row + 2, column)
                cell_updater.value = str(time)
                print("Logout time: " + cell_updater.value)
            else:
                current_sheet.cell(emp_row, column).value = "Present"
                cell_updater = current_sheet.cell(emp_row + 1, column)
                cell_updater.value = str(time)
                print("Login location: " + str(cell_updater))
                print("login time: " + cell_updater.value)


def find_emp(emp_name):
    current_sheet = current()

    for row in range(1, current_sheet.max_row):
        cell = current_sheet.cell(row, emp_column)
        if cell.value == emp_name:
            print("Emp location: " + str(cell))
            return cell.row


main("Nagaraju Peddarapu")
